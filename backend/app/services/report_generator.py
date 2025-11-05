"""
Report Generator for PM Document Intelligence
Generates PDF and Excel reports with analytics data
"""

from datetime import datetime, timedelta
from typing import Optional, Dict, Any, List
from sqlalchemy.orm import Session
from sqlalchemy import func
import io
import os
from pathlib import Path

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    PageBreak,
    Image,
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Excel generation
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill

# Charts for PDF
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
import base64

from app.services.analytics_service import AnalyticsService
from app.monitoring.log_aggregation import app_logger


class ReportGenerator:
    """Generate analytics reports in various formats"""

    def __init__(self, db: Session):
        self.db = db
        self.analytics_service = AnalyticsService(db)

    async def generate_and_send_report(
        self,
        report_type: str,
        start_date: datetime,
        end_date: datetime,
        format: str,
        email_to: str,
        user_id: str,
    ):
        """Generate report and send via email"""
        try:
            app_logger.info(
                f"Generating {report_type} report in {format} format", user_id=user_id
            )

            # Collect data
            report_data = await self._collect_report_data(start_date, end_date)

            # Generate report based on format
            if format == "pdf":
                report_file = await self.generate_pdf_report(
                    report_type, report_data, start_date, end_date
                )
            elif format == "excel":
                report_file = await self.generate_excel_report(
                    report_type, report_data, start_date, end_date
                )
            else:  # json
                report_file = await self.generate_json_report(
                    report_type, report_data, start_date, end_date
                )

            # Send email
            await self._send_report_email(email_to, report_type, report_file, format)

            app_logger.info(
                f"Report sent successfully to {email_to}",
                report_type=report_type,
                format=format,
            )

        except Exception as e:
            app_logger.error(
                f"Error generating/sending report: {str(e)}", exc_info=True
            )
            raise

    async def _collect_report_data(
        self, start_date: datetime, end_date: datetime
    ) -> Dict[str, Any]:
        """Collect all data needed for report"""
        from app.models.user import User
        from app.models.document import Document

        # Get admin user for full access
        admin_user = self.db.query(User).filter(User.is_superuser == True).first()

        if not admin_user:
            # Create temporary admin-like object
            class TempAdmin:
                is_superuser = True
                id = "system"

            admin_user = TempAdmin()

        # Document stats
        total_docs = (
            self.db.query(Document)
            .filter(Document.created_at.between(start_date, end_date))
            .count()
        )

        completed_docs = (
            self.db.query(Document)
            .filter(
                Document.created_at.between(start_date, end_date),
                Document.status == "completed",
            )
            .count()
        )

        # Get time series data
        time_series = self.analytics_service.get_document_time_series(
            start_date, end_date, None, admin_user
        )

        # Processing performance
        performance = self.analytics_service.get_processing_performance(
            start_date, end_date
        )

        # User activity
        active_users = (
            self.db.query(func.count(func.distinct(Document.user_id)))
            .filter(Document.created_at.between(start_date, end_date))
            .scalar()
        )

        # Cost data
        from app.monitoring.cost_tracking import cost_tracker

        total_cost = sum(
            entry.total_cost
            for entry in cost_tracker.cost_history
            if start_date <= entry.timestamp <= end_date
        )

        return {
            "total_documents": total_docs,
            "completed_documents": completed_docs,
            "success_rate": round(
                (completed_docs / total_docs * 100) if total_docs > 0 else 0, 2
            ),
            "time_series": time_series,
            "performance": performance,
            "active_users": active_users,
            "total_cost": total_cost,
            "date_range": {"start": start_date, "end": end_date},
        }

    async def generate_pdf_report(
        self,
        report_type: str,
        data: Dict[str, Any],
        start_date: datetime,
        end_date: datetime,
    ) -> bytes:
        """Generate PDF report with charts"""
        try:
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []

            # Title
            title_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Heading1"],
                fontSize=24,
                textColor=colors.HexColor("#1a1a1a"),
                spaceAfter=30,
                alignment=TA_CENTER,
            )

            title = Paragraph(
                f"PM Document Intelligence<br/>{report_type.title()} Report",
                title_style,
            )
            story.append(title)

            # Date range
            date_range = Paragraph(
                f"<b>Report Period:</b> {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}",
                styles["Normal"],
            )
            story.append(date_range)
            story.append(Spacer(1, 0.3 * inch))

            # Executive Summary
            story.append(Paragraph("<b>Executive Summary</b>", styles["Heading2"]))
            story.append(Spacer(1, 0.2 * inch))

            summary_data = [
                ["Metric", "Value"],
                ["Total Documents Processed", str(data["total_documents"])],
                ["Successfully Processed", str(data["completed_documents"])],
                ["Success Rate", f"{data['success_rate']}%"],
                ["Active Users", str(data["active_users"])],
                ["Total Cost", f"${data['total_cost']:.2f}"],
            ]

            summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
            summary_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 12),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ]
                )
            )

            story.append(summary_table)
            story.append(Spacer(1, 0.5 * inch))

            # Performance Metrics
            if data.get("performance"):
                story.append(
                    Paragraph("<b>Performance Metrics</b>", styles["Heading2"])
                )
                story.append(Spacer(1, 0.2 * inch))

                perf = data["performance"]["overall_metrics"]
                perf_data = [
                    ["Metric", "Value"],
                    ["Total Processed", str(perf["total_processed"])],
                    ["Completed", str(perf["completed"])],
                    ["Failed", str(perf["failed"])],
                    ["Success Rate", f"{perf['success_rate']}%"],
                ]

                perf_table = Table(perf_data, colWidths=[3 * inch, 2 * inch])
                perf_table.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("GRID", (0, 0), (-1, -1), 1, colors.black),
                        ]
                    )
                )

                story.append(perf_table)
                story.append(Spacer(1, 0.3 * inch))

            # Generate and add chart
            if data.get("time_series") and len(data["time_series"]) > 0:
                chart_image = self._create_time_series_chart(data["time_series"])
                if chart_image:
                    story.append(
                        Paragraph(
                            "<b>Document Processing Trend</b>", styles["Heading2"]
                        )
                    )
                    story.append(Spacer(1, 0.2 * inch))
                    story.append(chart_image)
                    story.append(Spacer(1, 0.3 * inch))

            # Footer
            story.append(PageBreak())
            footer = Paragraph(
                f"<i>Generated on {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC</i>",
                styles["Normal"],
            )
            story.append(footer)

            # Build PDF
            doc.build(story)

            buffer.seek(0)
            return buffer.read()

        except Exception as e:
            app_logger.error(f"Error generating PDF report: {str(e)}", exc_info=True)
            raise

    def _create_time_series_chart(
        self, time_series_data: List[Dict]
    ) -> Optional[Image]:
        """Create time series chart for PDF"""
        try:
            # Extract dates and counts
            dates = [entry["date"] for entry in time_series_data]
            counts = [entry["total"] for entry in time_series_data]

            # Create figure
            fig = Figure(figsize=(6, 4))
            ax = fig.add_subplot(111)

            ax.plot(dates, counts, marker="o", linestyle="-", linewidth=2, markersize=6)
            ax.set_xlabel("Date")
            ax.set_ylabel("Document Count")
            ax.set_title("Documents Processed Over Time")
            ax.grid(True, alpha=0.3)

            # Rotate x-axis labels
            plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha="right")

            # Save to buffer
            img_buffer = io.BytesIO()
            fig.savefig(img_buffer, format="png", bbox_inches="tight", dpi=100)
            img_buffer.seek(0)

            # Create ReportLab Image
            img = Image(img_buffer, width=5 * inch, height=3.5 * inch)

            return img

        except Exception as e:
            app_logger.error(f"Error creating chart: {str(e)}")
            return None

    async def generate_excel_report(
        self,
        report_type: str,
        data: Dict[str, Any],
        start_date: datetime,
        end_date: datetime,
    ) -> bytes:
        """Generate Excel report with charts"""
        try:
            buffer = io.BytesIO()
            wb = Workbook()

            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"

            # Title
            ws_summary["A1"] = (
                f"PM Document Intelligence - {report_type.title()} Report"
            )
            ws_summary["A1"].font = Font(size=16, bold=True)
            ws_summary["A2"] = (
                f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
            )

            # Summary data
            ws_summary["A4"] = "Metric"
            ws_summary["B4"] = "Value"
            ws_summary["A4"].font = Font(bold=True)
            ws_summary["B4"].font = Font(bold=True)

            summary_rows = [
                ("Total Documents", data["total_documents"]),
                ("Completed", data["completed_documents"]),
                ("Success Rate", f"{data['success_rate']}%"),
                ("Active Users", data["active_users"]),
                ("Total Cost", f"${data['total_cost']:.2f}"),
            ]

            for idx, (metric, value) in enumerate(summary_rows, start=5):
                ws_summary[f"A{idx}"] = metric
                ws_summary[f"B{idx}"] = value

            # Time Series sheet
            if data.get("time_series"):
                ws_timeseries = wb.create_sheet("Time Series")
                ws_timeseries["A1"] = "Date"
                ws_timeseries["B1"] = "Count"
                ws_timeseries["A1"].font = Font(bold=True)
                ws_timeseries["B1"].font = Font(bold=True)

                for idx, entry in enumerate(data["time_series"], start=2):
                    ws_timeseries[f"A{idx}"] = entry["date"]
                    ws_timeseries[f"B{idx}"] = entry["total"]

                # Add chart
                chart = LineChart()
                chart.title = "Documents Over Time"
                chart.y_axis.title = "Count"
                chart.x_axis.title = "Date"

                data_ref = Reference(
                    ws_timeseries,
                    min_col=2,
                    min_row=1,
                    max_row=len(data["time_series"]) + 1,
                )
                cats = Reference(
                    ws_timeseries,
                    min_col=1,
                    min_row=2,
                    max_row=len(data["time_series"]) + 1,
                )

                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats)

                ws_timeseries.add_chart(chart, "D2")

            # Performance sheet
            if data.get("performance"):
                ws_perf = wb.create_sheet("Performance")
                ws_perf["A1"] = "Performance Metrics"
                ws_perf["A1"].font = Font(size=14, bold=True)

                perf = data["performance"]["overall_metrics"]
                perf_rows = [
                    ("Metric", "Value"),
                    ("Total Processed", perf["total_processed"]),
                    ("Completed", perf["completed"]),
                    ("Failed", perf["failed"]),
                    ("Success Rate", f"{perf['success_rate']}%"),
                ]

                for idx, (metric, value) in enumerate(perf_rows, start=3):
                    ws_perf[f"A{idx}"] = metric
                    ws_perf[f"B{idx}"] = value
                    if idx == 3:
                        ws_perf[f"A{idx}"].font = Font(bold=True)
                        ws_perf[f"B{idx}"].font = Font(bold=True)

            # Save workbook
            wb.save(buffer)
            buffer.seek(0)

            return buffer.read()

        except Exception as e:
            app_logger.error(f"Error generating Excel report: {str(e)}", exc_info=True)
            raise

    async def generate_json_report(
        self,
        report_type: str,
        data: Dict[str, Any],
        start_date: datetime,
        end_date: datetime,
    ) -> bytes:
        """Generate JSON report"""
        try:
            import json

            report = {
                "report_type": report_type,
                "generated_at": datetime.utcnow().isoformat(),
                "date_range": {
                    "start": start_date.isoformat(),
                    "end": end_date.isoformat(),
                },
                "data": {
                    **data,
                    "date_range": {
                        "start": data["date_range"]["start"].isoformat(),
                        "end": data["date_range"]["end"].isoformat(),
                    },
                },
            }

            json_str = json.dumps(report, indent=2)
            return json_str.encode("utf-8")

        except Exception as e:
            app_logger.error(f"Error generating JSON report: {str(e)}", exc_info=True)
            raise

    async def _send_report_email(
        self, email_to: str, report_type: str, report_file: bytes, format: str
    ):
        """Send report via email"""
        try:
            # This would integrate with email service (SendGrid, SES, etc.)
            # For now, just log
            app_logger.info(
                f"Would send {format} report to {email_to}",
                report_type=report_type,
                file_size=len(report_file),
            )

            # Simulated email sending
            # import smtplib
            # from email.mime.multipart import MIMEMultipart
            # from email.mime.text import MIMEText
            # from email.mime.application import MIMEApplication
            #
            # msg = MIMEMultipart()
            # msg['From'] = 'noreply@pmdocintel.com'
            # msg['To'] = email_to
            # msg['Subject'] = f"PM Document Intelligence - {report_type.title()} Report"
            #
            # body = f"""
            # Hello,
            #
            # Please find attached your {report_type} analytics report.
            #
            # Best regards,
            # PM Document Intelligence Team
            # """
            #
            # msg.attach(MIMEText(body, 'plain'))
            #
            # attachment = MIMEApplication(report_file)
            # attachment['Content-Disposition'] = f'attachment; filename="{report_type}_report.{format}"'
            # msg.attach(attachment)
            #
            # # Send email via SMTP

        except Exception as e:
            app_logger.error(f"Error sending report email: {str(e)}", exc_info=True)
            raise
